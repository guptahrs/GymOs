import csv
import threading
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib.auth.hashers import make_password
from django.core.files.storage import default_storage
from django.db import close_old_connections, transaction
from django.utils import timezone

from accounts.models import User
from common.constants.enums import OnboardingStep, PaymentMode, PaymentStatus, SubscriptionStatus, UserType
from common.utills.plan_limits import ensure_member_capacity
from gyms.models import Gym
from members.models import (
    Member,
    MemberImportJob,
    MemberImportRowError,
    MemberPayment,
    MemberSubscription,
    MembershipPlan,
)

EXPECTED_COLUMNS = {
    "first_name",
    "last_name",
    "phone",
    "email",
    "gender",
    "dob",
    "start_date",
    "plan_type",
    "plan_name",
    "plan_price",
    "amount_paid",
    "pending_amount",
    "pending_due_date",
    "payment_mode",
}

REQUIRED_COLUMNS = {
    "first_name",
    "phone",
    "start_date",
    "plan_type",
    "plan_price",
    "amount_paid",
}

COLUMN_ALIASES = {
    "first name": "first_name",
    "firstname": "first_name",
    "last name": "last_name",
    "lastname": "last_name",
    "mobile": "phone",
    "mobile_number": "phone",
    "phone_number": "phone",
    "date_of_birth": "dob",
    "birth_date": "dob",
    "joining_date": "start_date",
    "subscription_start_date": "start_date",
    "membership_start_date": "start_date",
    "plan cycle": "plan_type",
    "plan_type": "plan_type",
    "billing_cycle": "plan_type",
    "plan": "plan_name",
    "membership_plan": "plan_name",
    "price": "plan_price",
    "plan_amount": "plan_price",
    "paid_amount": "amount_paid",
    "amount received": "amount_paid",
    "due_amount": "pending_amount",
    "remaining_amount": "pending_amount",
    "pending_due_date": "pending_due_date",
    "expected_date": "pending_due_date",
    "expected_payment_date": "pending_due_date",
    "payment_mode": "payment_mode",
    "payment_method": "payment_mode",
}

PLAN_TYPE_MAP = {
    "monthly": ("Monthly", 30),
    "month": ("Monthly", 30),
    "quarterly": ("Quarterly", 90),
    "quarter": ("Quarterly", 90),
    "3_month": ("Quarterly", 90),
    "3 months": ("Quarterly", 90),
    "yearly": ("Yearly", 365),
    "annual": ("Yearly", 365),
    "year": ("Yearly", 365),
}


def normalize_column_name(value):
    normalized = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return COLUMN_ALIASES.get(normalized, normalized)


def parse_upload_file(file_obj, file_name):
    suffix = (file_name or "").lower().rsplit(".", 1)[-1] if "." in (file_name or "") else ""
    if suffix == "csv":
        return _parse_csv(file_obj)
    if suffix == "xlsx":
        return _parse_xlsx(file_obj)
    raise ValueError("Only .xlsx and .csv files are supported.")


def validate_upload_columns(file_obj, file_name):
    rows = parse_upload_file(file_obj, file_name)
    if not rows:
        raise ValueError("The uploaded file is empty.")

    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    return len(rows)


def start_member_import(job_id):
    thread = threading.Thread(
        target=_process_member_import_job,
        args=(str(job_id),),
        daemon=True,
        name=f"member-import-{job_id}",
    )
    thread.start()


def _parse_csv(file_obj):
    file_obj.seek(0)
    decoded = file_obj.read().decode("utf-8-sig").splitlines()
    reader = csv.DictReader(decoded)
    rows = []
    for row in reader:
        normalized = {
            normalize_column_name(key): (value.strip() if isinstance(value, str) else value)
            for key, value in row.items()
            if key is not None
        }
        rows.append(normalized)
    return rows


def _parse_xlsx(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel support requires openpyxl. Please install backend requirements.") from exc

    file_obj.seek(0)
    workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    headers = [normalize_column_name(cell) for cell in header_row]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        row = {}
        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else ""
            if isinstance(value, datetime):
                value = value.date().isoformat()
            elif isinstance(value, date):
                value = value.isoformat()
            row[header] = "" if value is None else str(value).strip()
        rows.append(row)
    workbook.close()
    return rows


def _process_member_import_job(job_id):
    close_old_connections()
    job = MemberImportJob.objects.select_related("gym", "uploaded_by").get(import_job_id=job_id)
    job.status = MemberImportJob.STATUS_IN_PROGRESS
    job.started_at = timezone.now()
    job.summary_message = "Processing started"
    job.save(update_fields=["status", "started_at", "summary_message", "updated_at"])

    try:
        with default_storage.open(job.source_file.name, "rb") as source_file:
            rows = parse_upload_file(source_file, job.original_file_name)
    except Exception as exc:
        _fail_job(job, f"Unable to read file: {exc}")
        close_old_connections()
        return

    job.total_rows = len(rows)
    job.save(update_fields=["total_rows", "updated_at"])

    for index, raw_row in enumerate(rows, start=2):
        row_data = {key: (value if value is not None else "") for key, value in raw_row.items()}
        try:
            _process_single_row(job, row_data)
            job.success_rows += 1
        except Exception as exc:
            job.failed_rows += 1
            MemberImportRowError.objects.create(
                job=job,
                row_number=index,
                member_name=f"{row_data.get('first_name', '')} {row_data.get('last_name', '')}".strip(),
                phone=str(row_data.get("phone", "")).strip(),
                plan_name=str(row_data.get("plan_name") or row_data.get("plan_type") or "").strip(),
                error_message=str(exc),
                row_data=row_data,
            )
        finally:
            job.processed_rows += 1
            job.save(
                update_fields=[
                    "processed_rows",
                    "success_rows",
                    "failed_rows",
                    "updated_at",
                ]
            )

    job.completed_at = timezone.now()
    if job.failed_rows and job.success_rows:
        job.status = MemberImportJob.STATUS_COMPLETED_WITH_ERRORS
        job.summary_message = (
            f"Processed {job.processed_rows} rows. "
            f"{job.success_rows} succeeded and {job.failed_rows} failed."
        )
    elif job.failed_rows and not job.success_rows:
        job.status = MemberImportJob.STATUS_FAILED
        job.summary_message = f"All {job.failed_rows} rows failed during processing."
    else:
        job.status = MemberImportJob.STATUS_COMPLETED
        job.summary_message = f"All {job.success_rows} rows were imported successfully."
    job.save(
        update_fields=[
            "status",
            "completed_at",
            "summary_message",
            "updated_at",
        ]
    )
    close_old_connections()


def _process_single_row(job, row):
    gym = job.gym
    limit_error = ensure_member_capacity(gym.gym_id)
    if limit_error:
        raise ValueError(limit_error.data.get("message") or "Gym member limit reached.")

    normalized = _normalize_row(row)
    start_date = _parse_date(normalized["start_date"], field_name="start_date")
    pending_due_date = None
    if normalized.get("pending_due_date"):
        pending_due_date = _parse_date(normalized["pending_due_date"], field_name="pending_due_date")

    plan_name, duration_days = _resolve_plan_type(normalized.get("plan_type"), normalized.get("plan_name"))
    plan_price = _parse_decimal(normalized.get("plan_price"), "plan_price")
    amount_paid = _parse_decimal(normalized.get("amount_paid") or "0", "amount_paid")
    pending_amount = _parse_decimal(normalized.get("pending_amount") or "0", "pending_amount")

    if amount_paid < 0 or pending_amount < 0:
        raise ValueError("amount_paid and pending_amount cannot be negative.")
    if (amount_paid + pending_amount) != plan_price:
        raise ValueError("amount_paid + pending_amount must equal plan_price.")
    if pending_amount > 0 and not pending_due_date:
        raise ValueError("pending_due_date is required when pending_amount is greater than 0.")

    payment_mode = _normalize_payment_mode(normalized.get("payment_mode"))
    email = _resolve_email(gym, normalized.get("email"), normalized["phone"])

    if User.objects.filter(email=email, is_deleted=False).exists():
        raise ValueError(f"User with email '{email}' already exists.")
    if User.objects.filter(gym_id=gym.gym_id, phone=normalized["phone"], user_type=UserType.MEMBER, is_deleted=False).exists():
        raise ValueError(f"Member with phone '{normalized['phone']}' already exists for this gym.")

    with transaction.atomic():
        membership_plan, _ = MembershipPlan.objects.get_or_create(
            gym_id=gym.gym_id,
            name=plan_name,
            duration_days=duration_days,
            price=plan_price,
            defaults={"is_active": True, "is_deleted": False},
        )
        if not membership_plan.is_active or membership_plan.is_deleted:
            membership_plan.is_active = True
            membership_plan.is_deleted = False
            membership_plan.save(update_fields=["is_active", "is_deleted", "updated_at"])

        user = User.objects.create(
            first_name=normalized["first_name"],
            last_name=normalized.get("last_name", ""),
            email=email,
            phone=normalized["phone"],
            password=make_password("defaultpassword123"),
            user_type=UserType.MEMBER,
            gender=normalized.get("gender") or None,
            date_of_birth=_parse_optional_date(normalized.get("dob")),
            gym_id=gym.gym_id,
        )

        member = Member.objects.create(
            user=user,
            gym_id=gym.gym_id,
            date_of_birth=user.date_of_birth,
            onboarding_step=OnboardingStep.COMPLETED,
            payment_status=PaymentStatus.PAID if pending_amount == 0 else PaymentStatus.PARTIAL,
        )

        start_datetime = timezone.make_aware(datetime.combine(start_date, time.min))
        end_datetime = start_datetime + timedelta(days=membership_plan.duration_days)

        MemberSubscription.objects.create(
            gym_id=gym.gym_id,
            member=member,
            plan=membership_plan,
            start_date=start_datetime,
            end_date=end_datetime,
            amount_paid=amount_paid,
            remaining_amount=pending_amount,
            estimated_remaining_payment_date=pending_due_date,
            status=SubscriptionStatus.ACTIVE,
        )

        if amount_paid > 0:
            MemberPayment.objects.create(
                gym_id=gym.gym_id,
                member=member,
                amount=amount_paid,
                payment_mode=payment_mode,
            )


def _normalize_row(row):
    normalized = {}
    for key, value in row.items():
        normalized_key = normalize_column_name(key)
        normalized[normalized_key] = value.strip() if isinstance(value, str) else value

    missing = [column for column in REQUIRED_COLUMNS if not str(normalized.get(column, "")).strip()]
    if missing:
        raise ValueError(f"Missing required values: {', '.join(sorted(missing))}")

    return normalized


def _resolve_plan_type(plan_type, plan_name):
    normalized_plan_type = str(plan_type or "").strip().lower().replace("-", "_").replace(" ", "_")
    if normalized_plan_type not in PLAN_TYPE_MAP:
        raise ValueError("plan_type must be one of monthly, quarterly, or yearly.")
    default_name, duration_days = PLAN_TYPE_MAP[normalized_plan_type]
    final_name = str(plan_name or default_name).strip() or default_name
    return final_name, duration_days


def _parse_decimal(value, field_name):
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        raise ValueError(f"{field_name} must be a valid number.")


def _parse_date(value, field_name):
    raw = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name} must be a valid date in YYYY-MM-DD or DD/MM/YYYY format.")


def _parse_optional_date(value):
    if not value:
        return None
    return _parse_date(value, "dob")


def _normalize_payment_mode(value):
    raw = str(value or PaymentMode.CASH.value).strip().upper()
    if raw not in PaymentMode.values():
        raise ValueError("payment_mode must be one of CASH, CARD, or UPI.")
    return raw


def _resolve_email(gym, email, phone):
    cleaned_email = str(email or "").strip().lower()
    if cleaned_email:
        return cleaned_email

    safe_gym = str(gym.gym_id).replace("-", "")
    digits = "".join(ch for ch in str(phone or "") if ch.isdigit()) or "member"
    return f"import-{safe_gym[:8]}-{digits}@gymora.local"


def _fail_job(job, message):
    job.status = MemberImportJob.STATUS_FAILED
    job.completed_at = timezone.now()
    job.summary_message = message
    job.save(update_fields=["status", "completed_at", "summary_message", "updated_at"])
